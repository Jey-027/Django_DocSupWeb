from .files import genera_archivo, genera_archivo_usd, genera_nc
from .models import documento, proveedor, documento, properties
from django.views.generic.edit import CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, TemplateView
from django.contrib.auth.forms import UserCreationForm
from django.http.response import  HttpResponse
from django.shortcuts import redirect, render
from django.contrib.auth import logout
from django.core.mail import send_mail
from django.urls import reverse_lazy
from django.contrib import messages
from .forms import documentoForm, noFileForm, notaCredito, documentFormUsd
from django.conf import settings
from django.db.models import Q
from datetime import datetime
from openpyxl import Workbook
# import time

# Create your views here.


def logout_view(request):
    logout(request)
    return redirect("Home")


class homeView(LoginRequiredMixin, TemplateView):
    template_name = "DocSupApp/home.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["documento"] = documento.objects.all()
        context["documentoNull"] = documento.objects.filter(~Q(status=1))
        context["documento1"] = documento.objects.filter(status=1)
        context["notas"] = documento.objects.filter(genero_nota_credito=1)
        return context


class SignUp(CreateView):
    form_class = UserCreationForm
    success_url = reverse_lazy("login")
    template_name = "registration/signup.html"

class vendorList(LoginRequiredMixin,ListView):
    model = proveedor
    context_object_name = "proveedor_list"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["proveedores"] = proveedor.objects.all()
        return context

def search(request):
    template_name = "DocSupApp/proveedor_list.html"
    queryset = request.GET["Search"]
    vendor = proveedor.objects.filter(name__icontains = queryset)
    context = {
        'vendor': vendor
    }

    return render(request, template_name, context)

class VendorCreate(LoginRequiredMixin,CreateView):
    model = proveedor
    template_name = "DocSupApp/proveedor_create.html"
    fields = "__all__"
    success_url = "/vendor/list"

class VendorUpdate(LoginRequiredMixin,UpdateView):
    model = proveedor
    template_name = "DocSupApp/proveedor_update.html"
    fields = "__all__"
    success_url = "/vendor/list"

# aqui van las view sobre la generacion del documento

class DetFactList(LoginRequiredMixin,ListView):
    model = documento
    context_object_name = "lista_de_Documentos"

class detalle_documentos(LoginRequiredMixin, ListView):
    model = documento
    context_object_name = "lista_de_Documentos2"
    template_name = "DocSupApp/notas_credito.html"


def noSendFile(request, id):
    doc2 = documento.objects.get(id = id)

    if request.method == "GET":
        form = noFileForm(instance=doc2)
    else:
        form = noFileForm(request.POST, instance=doc2)

        if form.is_valid():
            doc2.Date_process = datetime.now()
            doc2.status = 2
            doc2.user_process = str(request.user)
            form.save()
            messages.add_message(request=request, level=messages.SUCCESS, message="Documento no enviado, pero actualizado!")
       
        return redirect("Detalle_facturacion")
    return render(request, "DocSupApp/generacion_documento.html", {"form": form})


def genera_notaCredito(request,id):
    doc3 = documento.objects.get(id = id)
    num_res = properties.objects.get(id = 2)

    if request.method == "GET":
        form = notaCredito(instance=doc3)
    else:
        form = notaCredito(request.POST, instance=doc3)
        
        if form.is_valid():
            doc3.genero_nota_credito = "True"
            doc3.fecha_NC = datetime.now()
            doc3.user_create_NC = str(request.user)
            doc3.name_file_NC = doc3.num_documento + "-NC"
            form.save()

            genera_nc(id)

        num_res.Num_resolution += 1
        num_res.save()
        messages.add_message(request=request, level=messages.SUCCESS, message="Nota credito creada correctamente!")

        return redirect("Doc_generadas")
    return render(request, "DocSupApp/genera_notaCredito.html", {"form": form}) 


def updateDocumento(request, id):
    doc = documento.objects.get(id = id)
    numRes = properties.objects.first()

    if request.method == "GET" and doc.type_of_tax_number != '13':
        form = documentFormUsd(instance=doc)
    elif request.method == "GET" and doc.type_of_tax_number == '13':
        form = documentoForm(instance=doc)
    else:
        if doc.type_of_tax_number == '13':
            form = documentoForm(request.POST, instance=doc)
            if form.is_valid():
                doc.num_documento = numRes.name_file + str(numRes.Num_resolution)
                doc.Date_process = datetime.now()
                doc.status = 1
                doc.user_process = str(request.user)
                form.save()

        else:
            form = documentFormUsd(request.POST, instance=doc)
            if form.is_valid():
                doc.num_documento = numRes.name_file + str(numRes.Num_resolution)
                doc.Date_process = datetime.now()
                doc.status = 1
                doc.user_process = str(request.user)
                # form.save()

        try:
            if doc.type_of_tax_number == '13':
                genera_archivo(id)
                form.save()
                numRes.Num_resolution += 1
                numRes.save()
                messages.add_message(request=request, level=messages.SUCCESS, message="Documento generado correctamente!")
            else:
                genera_archivo_usd(id)
                form.save()
                numRes.Num_resolution += 1
                numRes.save()
                messages.add_message(request=request, level=messages.SUCCESS, message="Documento generado correctamente!")
        except:
            doc.status = 'NULL'
            form.save()
        messages.add_message(request=request, level=messages.ERROR, message="Documento no generado!")


        if numRes.Num_resolution >= 5000030:
            subject = "advertencia numero consecutivo doc soporte"
            message = "este correo es para notificar que su numero de resolucion es " + str(numRes.Num_resolution) + "para que vaya solicitando su nuevo rango"
            email_from = settings.EMAIL_HOST_USER
            recipient_list=["jdrodriguezg25@hotmail.com"]
            send_mail(subject, message, email_from, recipient_list)

            

        return redirect("Detalle_facturacion")
    return render(request, "DocSupApp/generacion_documento.html", {"form": form})




class reporteExcelVendor(TemplateView):
    def get(self, request, *args, **kwargs):
        vendor = proveedor.objects.all()
        wb = Workbook()
        ws = wb.active
        ws["A1"]= "REPORTE DE PROVEEDORES"

        ws.merge_cells("A1:N1")
        ws["A3"]= "Id Supplier"
        ws["B3"]= "Name"
        ws["C3"]= "Tax code"
        ws["D3"]= "city id"
        ws["E3"]= "City Name"
        ws["F3"]= "Email"
        ws["G3"]= "Nit"
        ws["H3"]= "Tax Description"
        ws["I3"]= "Type of tax number"
        ws["J3"]= "Address"
        ws["K3"]= "Country"
        ws["L3"]= "Estado federal"
        ws["M3"]= "name estado federal"
        ws["N3"]= "Currency type"

        cont = 4

        for vendor in vendor:
            ws.cell(row= cont, column= 1).value = vendor.id_supplier
            ws.cell(row= cont, column= 2).value = vendor.name
            ws.cell(row= cont, column= 3).value = vendor.supplier_tax_code
            ws.cell(row= cont, column= 4).value = vendor.city_id
            ws.cell(row= cont, column= 5).value = vendor.city_name
            ws.cell(row= cont, column= 6).value = vendor.email
            ws.cell(row= cont, column= 7).value = vendor.nit
            ws.cell(row= cont, column= 8).value = vendor.supplier_tax_description
            ws.cell(row= cont, column= 9).value = vendor.Type_of_tax_number
            ws.cell(row= cont, column= 10).value = vendor.address
            ws.cell(row= cont, column= 11).value = vendor.country
            ws.cell(row= cont, column= 12).value = vendor.est_fed_prov
            ws.cell(row= cont, column= 13).value = vendor.name_est_fed_prov
            ws.cell(row= cont, column= 14).value = vendor.currency_type
            cont += 1

        file_name = "ReporteProveedores.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response

class reporteListaDocumentos(TemplateView):
    def get(self,request, *args, **kwargs):
        document = documento.objects.all()
        wb = Workbook()
        ws = wb.active
        ws["A1"]= "REPORTE DE DOCUMENTOS "

        ws.merge_cells("A1:N1")
        ws["A3"]= "Id supplier vendor"
        ws["B3"]= "Name vendor"
        ws["C3"]= "Tax code"
        ws["D3"]= "City Name"
        ws["E3"]= "Nit"
        ws["F3"]= "Type tax number"
        ws["G3"]= "Currency type"
        ws["H3"]= "Supplier id"
        ws["I3"]= "Name supplier customer"
        ws["J3"]= "COFP"
        ws["K3"]= "Date invoice"
        ws["L3"]= "Item description"
        ws["M3"]= "ZsupplierID"
        ws["N3"]= "zSupplierName"
        ws["O3"]= "Net Amount"
        ws["P3"]= "Tax Amount"
        ws["Q3"]= "Gross Amount"
        ws["R3"]= "Percent RTE"
        ws["S3"]= "Percent IVA"
        ws["T3"]= "Percent ICA"
        ws["U3"]= "Amount IVA"
        ws["V3"]= "Amount RTE"
        ws["W3"]= "Amount ICA"
        ws["X3"]= "Value RTE"
        ws["Y3"]= "Total Retenciones"
        ws["Z3"]= "Person Type"
        ws["AA3"]= "Status"
        ws["AB3"]= "Date Process"
        ws["AC3"]= "User Process"
        ws["AD3"]= "Payment date"
        ws["AE3"]= "Document number"

        cont = 4

        for document in document:
            ws.cell(row= cont, column=1).value = document.id_supplier_vendor
            ws.cell(row= cont, column=2).value = document.name_supplier_vendor
            ws.cell(row= cont, column=3).value = document.suplier_tax_code
            ws.cell(row= cont, column=4).value = document.city_name
            ws.cell(row= cont, column=5).value = document.Nit
            ws.cell(row= cont, column=6).value = document.type_of_tax_number
            ws.cell(row= cont, column=7).value = document.currency_type
            ws.cell(row= cont, column=8).value = document.supplier_id_Customer
            ws.cell(row= cont, column=9).value = document.name_supplier_customer
            ws.cell(row= cont, column=10).value = document.id_supplier_invoice
            ws.cell(row= cont, column=11).value = document.date_Invoice
            ws.cell(row= cont, column=12).value = document.item_description
            ws.cell(row= cont, column=13).value = document.zSupplierID
            ws.cell(row= cont, column=14).value = document.zSupplierName
            ws.cell(row= cont, column=15).value = document.net_amount
            ws.cell(row= cont, column=16).value = document.tax_amount
            ws.cell(row= cont, column=17).value = document.gross_amount
            ws.cell(row= cont, column=18).value = document.percent_RTE
            ws.cell(row= cont, column=19).value = document.percent_IVA
            ws.cell(row= cont, column=20).value = document.percent_ICA
            ws.cell(row= cont, column=21).value = document.amount_RTE
            ws.cell(row= cont, column=22).value = document.amount_IVA
            ws.cell(row= cont, column=23).value = document.amount_ICA
            ws.cell(row= cont, column=24).value = document.Value_RTE
            ws.cell(row= cont, column=25).value = document.total_retenciones
            ws.cell(row= cont, column=26).value = document.tipo_persona
            ws.cell(row= cont, column=27).value = document.status
            ws.cell(row= cont, column=28).value = document.Date_process
            ws.cell(row= cont, column=29).value = document.user_process
            ws.cell(row= cont, column=30).value = document.payment_date
            ws.cell(row= cont, column=31).value = document.num_documento
            cont +=1

        file_name = "ReporteDocumentos.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response



  


